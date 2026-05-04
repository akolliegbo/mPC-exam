"""
Per-patient fitting of PK/PD model to the 17-patient mCRPC adaptive therapy dataset.

Model structure (Brady-Nicholls & Enderling 2020, Nature Communications):
    State vector: [S, R, P]
        S  : differentiated (drug-sensitive) PCa cells  — paper's 'D'
        R  : stem-like (drug-resistant) PCa cells       — paper's 'S'
        P  : serum PSA (ng/mL)

    ODEs:
        dR/dt = (R/(R+S)) * p_s * λ * R
        dS/dt = (1 - R/(R+S) * p_s) * λ * R  -  α(C) * S
        dP/dt = δ * S  -  γ * P

    where α(C) = alpha_max * C / (EC50 + C)   [Emax PD; our extension]

Fixed parameters:
    λ   = ln(2)  day^-1   stem cell proliferation rate (one division/day)
    EC50 = 15    ng/mL    half-maximal drug concentration
    n_hill = 1             Hill coefficient

Fitted per patient:
    S0        initial differentiated cell fraction
    R0        initial stem-like cell fraction
    p_s       stem cell self-renewal probability
    alpha_max maximum drug kill rate (day^-1)
    gamma     PSA clearance rate (day^-1)

delta is derived analytically so PSA(0) = observed PSA(0):
    δ = psa0 * γ / S0

PK approximation (QSS):
    C = C_ss_avg ≈ 26.9 ng/mL when Abi=1, 0 when Abi=0.
    C_ss_avg = dose / (ke * V) * 1e3

Reference:
    Brady-Nicholls & Enderling (2020) Nat Commun; Stuyckens et al. (2014)
"""

import numpy as np
from scipy.integrate import solve_ivp
from scipy.optimize import differential_evolution
import openpyxl
import matplotlib.pyplot as plt
import matplotlib
import warnings
warnings.filterwarnings('ignore')

matplotlib.rcParams.update({'font.family': 'Arial', 'font.size': 12})


# ── Fixed parameters ───────────────────────────────────────────────────────────

P_BASE = {
    # Tumor dynamics — Brady-Nicholls 2020
    'lambda_': np.log(2),  # day^-1   stem cell proliferation rate (1 division/day)
    # PK — Stuyckens et al. 2014, mCRPC population means, 1000 mg QD fasted
    'ka':      39.6,       # day^-1   (not used in QSS fitting; retained for reference)
    'ke':      1.386,      # day^-1   elimination rate constant (from t½ = 12 h)
    'V':       26840.0,    # L        apparent V/F
    'dose_mg': 1000.0,     # mg       abiraterone acetate QD fasted
    # PD — Emax model; EC50 fixed, alpha_max fitted per patient
    'EC50':    15.0,       # ng/mL    fixed (~0.5 * C_ss_avg)
    'n_hill':  1.0,        # Hill coefficient (fixed)
}

# Quasi-steady-state concentration (derived from PK parameters)
C_SS = P_BASE['dose_mg'] / (P_BASE['ke'] * P_BASE['V']) * 1e3  # ≈ 26.9 ng/mL


# ── Data loading ───────────────────────────────────────────────────────────────

def load_adaptive_patients(filepath):
    """
    Load the adaptive therapy arm from TrialPatientData.xlsx.

    Returns
    -------
    dict: {patient_id: {'days': array, 'psa': array, 'abi': array}}
    """
    wb = openpyxl.load_workbook(filepath)
    ws = wb['Adaptive']

    row2 = [ws.cell(2, c).value for c in range(1, ws.max_column + 1)]
    row1 = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    day_col_indices = [i for i, v in enumerate(row2) if v == 'Days']

    patients = {}
    for col_idx in day_col_indices:
        pid = row1[col_idx]
        days, psa, abi = [], [], []
        for r in range(3, ws.max_row + 1):
            d   = ws.cell(r, col_idx + 1).value
            p_v = ws.cell(r, col_idx + 2).value
            a   = ws.cell(r, col_idx + 3).value
            if d is None:
                break
            days.append(float(d))
            psa.append(float(p_v))
            abi.append(int(a))
        patients[pid] = {
            'days': np.array(days),
            'psa':  np.array(psa),
            'abi':  np.array(abi),
        }
    return patients


# ── ODE system (state: [S, R, P]) ─────────────────────────────────────────────

def odes(t, y, p, C, delta):
    """
    Brady-Nicholls 2020 stem cell model with Emax PD extension.

    S = differentiated (sensitive) cells  [paper's D]
    R = stem-like (resistant) cells       [paper's S]

    dR/dt = (R/(R+S)) * p_s * λ * R
    dS/dt = (1 - R/(R+S) * p_s) * λ * R  -  α(C) * S
    dP/dt = δ * S  -  γ * P

    Parameters
    ----------
    C     : float  drug concentration during this interval (ng/mL)
    delta : float  PSA production rate (derived from initial conditions)
    """
    S, R, P = y

    C_pos = max(C, 0.0)
    alpha = (p['alpha_max'] * C_pos**p['n_hill']
             / (p['EC50']**p['n_hill'] + C_pos**p['n_hill']))

    N = S + R
    stem_frac = R / N if N > 1e-12 else 0.0

    dR = stem_frac * p['p_s'] * p['lambda_'] * R
    dS = (1.0 - stem_frac * p['p_s']) * p['lambda_'] * R - alpha * S
    dP = delta * S - p['gamma'] * P

    return [dS, dR, dP]


# ── Single-patient simulation ──────────────────────────────────────────────────

def simulate_patient(p, days_obs, abi_obs, S0, R0, alpha_max, psa0):
    """
    Simulate model for one patient using the QSS PK approximation.

    Integrates between consecutive observation timepoints; C is set to
    C_ss_avg (26.9 ng/mL) when Abi=1, and 0 when Abi=0.

    Returns
    -------
    psa_sim : array of simulated PSA at days_obs (length = len(days_obs))
    """
    p = p.copy()
    p['alpha_max'] = alpha_max

    # Derive delta so PSA(0) = psa0 exactly (only S produces PSA)
    delta = psa0 * p['gamma'] / S0

    y = np.array([S0, R0, psa0])
    psa_sim = [psa0]

    for i in range(len(days_obs) - 1):
        C = C_SS if abi_obs[i] == 1 else 0.0

        sol = solve_ivp(
            fun=lambda t, y: odes(t, y, p, C, delta),
            t_span=(days_obs[i], days_obs[i + 1]),
            y0=y,
            method='RK45',
            rtol=1e-7,
            atol=1e-10,
        )
        y = sol.y[:, -1].copy()
        y[0] = max(y[0], 1e-9)   # keep S non-negative
        y[1] = max(y[1], 1e-9)   # keep R non-negative

        psa_sim.append(float(y[2]))

    return np.array(psa_sim)


# ── Objective function ─────────────────────────────────────────────────────────

def objective(params, days_obs, psa_obs, abi_obs, p_base):
    S0, R0, p_s, alpha_max, gamma = params

    if S0 <= 0 or R0 <= 0 or p_s <= 0 or p_s >= 1 or alpha_max <= 0 or gamma <= 0:
        return 1e6

    p = p_base.copy()
    p['p_s']   = p_s
    p['gamma'] = gamma

    try:
        psa_sim = simulate_patient(
            p, days_obs, abi_obs, S0, R0, alpha_max, psa_obs[0]
        )
        psa_sim = np.maximum(psa_sim, 1e-6)
        return float(np.sum((np.log(psa_sim) - np.log(psa_obs)) ** 2))
    except Exception:
        return 1e6


# ── Per-patient fitting ────────────────────────────────────────────────────────

def fit_patient(pid, data, p_base):
    """
    Fit S0, R0, p_s, alpha_max, gamma for one patient using differential evolution.

    Returns
    -------
    dict with fitted parameters and simulation output
    """
    days_obs = data['days']
    psa_obs  = data['psa']
    abi_obs  = data['abi']

    # p_s ∈ (0, 0.5): self-renewal probability; >0.5 would expand the stem pool
    # unboundedly. Brady-Nicholls median p_s = 0.0278.
    # gamma ∈ [0.05, 1.0] day^-1: PSA clearance. Brady-Nicholls phi = 0.0856.
    bounds = [
        (0.01,  0.99),   # S0        initial differentiated fraction
        (0.001, 0.50),   # R0        initial stem-like fraction
        (0.001, 0.499),  # p_s       stem self-renewal probability
        (0.001, 0.50),   # alpha_max max drug kill rate (day^-1)
        (0.05,  1.00),   # gamma     PSA clearance rate (day^-1)
    ]

    result = differential_evolution(
        objective,
        bounds=bounds,
        args=(days_obs, psa_obs, abi_obs, p_base),
        seed=42,
        maxiter=300,
        popsize=15,
        tol=1e-4,
        mutation=(0.5, 1.5),
        recombination=0.9,
        polish=False,
        workers=1,
    )

    S0, R0, p_s, alpha_max, gamma = result.x
    p_fit = p_base.copy()
    p_fit['p_s']   = p_s
    p_fit['gamma'] = gamma
    psa_sim = simulate_patient(
        p_fit, days_obs, abi_obs, S0, R0, alpha_max, psa_obs[0]
    )

    return {
        'pid':       pid,
        'S0':        S0,
        'R0':        R0,
        'p_s':       p_s,
        'alpha_max': alpha_max,
        'gamma':     gamma,
        'days':      days_obs,
        'psa_obs':   psa_obs,
        'psa_sim':   psa_sim,
        'abi':       abi_obs,
        'fun':       result.fun,
    }


# ── Plotting ───────────────────────────────────────────────────────────────────

def plot_fits(fits, ncols=4):
    n     = len(fits)
    nrows = int(np.ceil(n / ncols))
    fig, axes = plt.subplots(nrows, ncols, figsize=(ncols * 4, nrows * 3))
    axes = axes.flatten()

    for ax, fit in zip(axes, fits):
        t_yr = fit['days'] / 365.0

        # Shade treatment-ON periods
        for i in range(len(fit['days']) - 1):
            if fit['abi'][i] == 1:
                ax.axvspan(fit['days'][i] / 365, fit['days'][i + 1] / 365,
                           color='steelblue', alpha=0.12, lw=0)

        ax.semilogy(t_yr, fit['psa_obs'], 'o', color='darkorange', ms=4,
                    label='Observed')
        ax.semilogy(t_yr, fit['psa_sim'], '-', color='steelblue', lw=1.5,
                    label='Fitted')
        ax.set_title(f"{fit['pid']}\n"
                     f"S0={fit['S0']:.2f}, R0={fit['R0']:.3f}, "
                     f"p_s={fit['p_s']:.4f}, α_max={fit['alpha_max']:.3f}",
                     fontsize=9)
        ax.set_xlabel('Time (years)', fontsize=8)
        ax.set_ylabel('PSA (ng/mL)', fontsize=8)
        ax.spines[['top', 'right']].set_visible(False)

    for ax in axes[n:]:
        ax.set_visible(False)

    axes[0].legend(frameon=False, fontsize=8)
    plt.suptitle('Per-patient PK/PD fits — adaptive therapy arm\n'
                 f'Fixed: λ=ln(2), EC50={P_BASE["EC50"]} ng/mL, '
                 f'C_ss={C_SS:.1f} ng/mL',
                 fontsize=12, y=1.01)
    plt.tight_layout()
    return fig


def print_summary(fits):
    alpha_eff_factor = C_SS / (P_BASE['EC50'] + C_SS)
    print(f"\n{'Patient':<10} {'S0':>6} {'R0':>7} {'p_s':>8} "
          f"{'alpha_max':>10} {'alpha_eff':>10} {'gamma':>7} {'SSE_log':>8}")
    print('-' * 75)
    for f in fits:
        alpha_eff = f['alpha_max'] * alpha_eff_factor
        print(f"{f['pid']:<10} {f['S0']:>6.3f} {f['R0']:>7.4f} "
              f"{f['p_s']:>8.4f} {f['alpha_max']:>10.4f} "
              f"{alpha_eff:>10.4f} {f['gamma']:>7.4f} {f['fun']:>8.3f}")
    print(f"\nC_ss_avg = {C_SS:.1f} ng/mL  |  "
          f"alpha_eff = alpha_max x {alpha_eff_factor:.3f}  (EC50={P_BASE['EC50']} ng/mL)")


# ── Run ────────────────────────────────────────────────────────────────────────

if __name__ == '__main__':
    DATA_PATH = ('/Users/80031987/Desktop/Basanta_Marusyk_Lab_2026/'
                 'PKPD_AdaptiveTherapy/TrialPatientData.xlsx')

    print("Loading patient data...")
    patients = load_adaptive_patients(DATA_PATH)
    print(f"  {len(patients)} patients: {list(patients.keys())}\n")

    fits = []
    for i, (pid, data) in enumerate(patients.items()):
        print(f"[{i+1:2d}/{len(patients)}] Fitting {pid}...", end=' ', flush=True)
        fit = fit_patient(pid, data, P_BASE)
        fits.append(fit)
        print(f"S0={fit['S0']:.3f}, R0={fit['R0']:.4f}, p_s={fit['p_s']:.4f}, "
              f"alpha_max={fit['alpha_max']:.4f}, SSE={fit['fun']:.3f}")

    print_summary(fits)

    fig = plot_fits(fits)
    plt.tight_layout()
    plt.savefig('/Users/80031987/Desktop/Basanta_Marusyk_Lab_2026/'
                'PKPD_AdaptiveTherapy/patient_fits.pdf',
                dpi=300, bbox_inches='tight')
    print("\nFigure saved to patient_fits.pdf")
    plt.show()
